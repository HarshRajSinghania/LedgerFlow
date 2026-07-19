"""Writes the audit-ready Excel report: Summary, Reconciliation Report (all
lines), and Exceptions (action queue, sorted by financial risk)."""

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"

STATUS_FILL = {
    "MATCHED": PatternFill("solid", fgColor="C6EFCE"),
    "MISSING_IN_INVOICE": PatternFill("solid", fgColor="FFEB9C"),
    "MISSING_IN_PO": PatternFill("solid", fgColor="FFC7CE"),
    "VALUE_MISMATCH": PatternFill("solid", fgColor="FFC7CE"),
    "DUPLICATE_KEY_IN_PO": PatternFill("solid", fgColor="D9D2E9"),
    "DUPLICATE_KEY_IN_INVOICE": PatternFill("solid", fgColor="D9D2E9"),
}
STATUS_FONT_COLOR = {
    "MATCHED": "006100",
    "MISSING_IN_INVOICE": "9C6500",
    "MISSING_IN_PO": "9C0006",
    "VALUE_MISMATCH": "9C0006",
    "DUPLICATE_KEY_IN_PO": "4B0082",
    "DUPLICATE_KEY_IN_INVOICE": "4B0082",
}

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REPORT_COLUMNS = [
    "Invoice_ID", "Product_Code", "Status", "Supplier", "Product_Name",
    "PO_Date", "Invoice_Date", "PO_Quantity", "Invoice_Quantity",
    "PO_Unit_Price", "Invoice_Unit_Price", "PO_Total", "Invoice_Total",
    "Financial_Exposure", "Risk_Score", "Risk_Flags", "AI_Confidence", "AI_Note",
    "Action", "Action_Note",
]


def _prepare_for_excel(report: pd.DataFrame) -> pd.DataFrame:
    """Fills in any optional columns that may not exist yet (Risk_*, Action*
    are added by later pipeline stages) and makes list-typed cells
    (Risk_Flags) writable as plain text."""
    df = report.copy()
    for col in REPORT_COLUMNS:
        if col not in df.columns:
            df[col] = "" if col != "Risk_Score" else 0
    df["Risk_Flags"] = df["Risk_Flags"].apply(
        lambda v: ", ".join(v) if isinstance(v, list) else (v or "")
    )
    df["Action"] = df["Action"].fillna("")
    df["Action_Note"] = df["Action_Note"].fillna("")
    return df


def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autosize(ws, df, max_width=45):
    for i, col in enumerate(df.columns, start=1):
        width = min(max_width, max(12, int(df[col].astype(str).str.len().quantile(0.9)) + 4, len(str(col)) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_table(ws, df: pd.DataFrame, status_col="Status"):
    ws.append(list(df.columns))
    _style_header(ws, len(df.columns))

    for _, row in df.iterrows():
        ws.append(list(row))
        r = ws.max_row
        status = row.get(status_col, "")
        fill = STATUS_FILL.get(status)
        font_color = STATUS_FONT_COLOR.get(status, "000000")
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, color=font_color if fill else "000000")
            if fill:
                cell.fill = fill
            header = df.columns[c - 1]
            if "Price" in header or "Total" in header or "Exposure" in header:
                cell.number_format = '\"£\"#,##0.00;[RED]-\"£\"#,##0.00'
            if "Date" in header and header not in ("Product_Name",):
                cell.number_format = "dd/mm/yyyy"

    ws.freeze_panes = "A2"
    _autosize(ws, df)


def write_report(path, report: pd.DataFrame, stats: dict, summary_text: str, warnings: list):
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "LedgerFlow Invoice Reconciliation Report"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=16)
    ws["A2"] = f"Generated {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name=FONT_NAME, italic=True, size=10, color="666666")

    labels = [
        ("Total PO-Product lines reconciled", stats["total_lines"]),
        ("Matched (cleared for payment)", stats["matched"]),
        ("Missing in Invoice (ordered, not yet billed)", stats["missing_in_invoice"]),
        ("Missing in PO (billed, no matching order)", stats["missing_in_po"]),
        ("Value mismatches (billed differently than ordered)", stats["value_mismatch"]),
        ("Duplicate keys (manual review needed)", stats["duplicates"]),
        ("Total financial exposure in unresolved lines", f"£{stats['total_exposure']:,.2f}"),
        ("PO file row count (after cleaning)", stats["po_row_count"]),
        ("Invoice file row count (after cleaning)", stats["invoice_row_count"]),
    ]
    r = 4
    for label, value in labels:
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, bold=True)
        ws.cell(row=r, column=2, value=value).font = Font(name=FONT_NAME)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Executive Summary").font = Font(name=FONT_NAME, bold=True, size=12)
    r += 1
    cell = ws.cell(row=r, column=1, value=summary_text)
    cell.font = Font(name=FONT_NAME)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 6, end_column=6)
    ws.row_dimensions[r].height = 100
    r += 8

    if warnings:
        ws.cell(row=r, column=1, value="Ingestion Warnings (data quality issues found automatically)").font = \
            Font(name=FONT_NAME, bold=True, size=12, color="9C0006")
        r += 1
        for w in warnings:
            cell = ws.cell(row=r, column=1, value=f"- {w}")
            cell.font = Font(name=FONT_NAME, color="9C0006")
            cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            r += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 22

    # --- Full reconciliation report ---
    prepared = _prepare_for_excel(report)
    ws2 = wb.create_sheet("Reconciliation Report")
    full = prepared[REPORT_COLUMNS].copy()
    _write_table(ws2, full)

    # --- Exceptions only, sorted by financial risk (already sorted upstream) ---
    ws3 = wb.create_sheet("Exceptions")
    exceptions = prepared[prepared["Status"] != "MATCHED"][REPORT_COLUMNS].copy()
    _write_table(ws3, exceptions)

    wb.save(path)
    return path